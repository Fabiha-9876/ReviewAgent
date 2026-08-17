"""
Build Round-2 annotation materials for the 490-review gold-standard verification.

Why this exists
---------------
Round 1 produced exactly one usable filled sheet. The A/B/C workbooks on disk carry
byte-identical annotation columns, so no genuine inter-annotator agreement can be
computed from them. Round 2 re-runs the same task with two additional independent
human raters so Fleiss kappa and Krippendorff alpha are computed from real
independent judgements.

Design
------
- Same 490 reviews as Round 1, taken from annotator_materials/master_key.json
  ("main_indices"). Agreement statistics require the identical item set.
- Row order is the Round-1 order, identical across annotator_A, D and E, so the
  three sheets line up row by row for manual side-by-side cross-checking. This
  trades away order-effect control in favour of reviewability; alignment in code
  still goes through the row_id column, never row position.
- Annotation columns ship blank. Nothing is pre-filled.
- The same calibration set of 20 reviews is included, to be completed first.

Outputs (annotator_materials/round2/)
    00_README.md
    calibration_set_round2.xlsx
    annotator_D.xlsx
    annotator_E.xlsx
    round2_key.json          row order per annotator; keep private

Usage
    python3 scripts/build_annotator_round2.py
"""

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

LABELS = ["bug_report", "feature_request", "performance", "usability",
          "compatibility", "praise", "other"]

REPO = Path(__file__).resolve().parent.parent
DATASET = REPO / "data/processed/rrgen_v5_relabeled/rrgen_v5_relabeled.json"
MASTER_KEY = REPO / "annotator_materials/master_key.json"
OUT_DIR = REPO / "annotator_materials/round2"

# annotator code -> display name. Row order is the Round-1 order for everyone.
ANNOTATORS = {
    "D": "Annotator D",
    "E": "Annotator E",
}

HEADERS = ["row_id", "review_text", "rating", "app_id",
           "predicted_label", "correct_yn", "correct_label_if_no", "comments"]

INSTRUCTIONS = [
    "REVIEW VERIFICATION TASK",
    "",
    "Total reviews to verify: 490",
    "Estimated time: ~5-10 seconds per review",
    "",
    "What to do:",
    "  1. Read the review text in column 'review_text'.",
    "  2. Look at column 'predicted_label' - that's the AI's guess.",
    "  3. In column 'correct_yn', type Y if the AI label is correct, N if wrong.",
    "  4. If N, type the correct label in 'correct_label_if_no'.",
    "  5. Optional: add a one-line note in 'comments' for ambiguous cases.",
    "",
    "Use these exact label strings (lowercase, underscore):",
    "  bug_report        Crashes, errors, broken features",
    "  feature_request   Requests for new features or improvements",
    "  performance       Speed, battery, memory, lag, loading times",
    "  usability         Confusing UI, hard to use, poor navigation",
    "  compatibility     Device-specific or OS-specific issues",
    "  praise            Positive feedback, compliments",
    "  other             Doesn't fit any above category",
    "",
    "Decision rules (read these - they prevent the most common mistakes):",
    "  - 'slow', 'lag' is performance, NOT bug_report.",
    "  - 'crashes on my Samsung' is compatibility (device-specific).",
    "  - 'crash' alone (no device) is bug_report.",
    "  - 'would be nice if X' or 'please add X' is feature_request.",
    "  - 'hard to find the X button' is usability, not bug_report.",
    "  - Multi-aspect reviews: pick the primary category. If genuinely two-headed,",
    "    mark Y if the AI picked one of them.",
    "  - Spam / non-English / nonsense is other.",
    "  - When in doubt, mark Y if the label is reasonable; mark N only if",
    "    you're confident the label is wrong.",
    "",
    "Do NOT consult the other annotator or any previous annotation sheet.",
    "Independent judgements are the entire point of this round.",
    "",
    "Save the file and send it back when done. Resume any time - your fills",
    "are preserved in the cells.",
    "",
]


def load_selected_records():
    """Return {row_id: record} for the 490 Round-1 indices."""
    key = json.loads(MASTER_KEY.read_text())
    indices = [int(i) for i in key["main_indices"]]
    wanted = set(indices)

    print(f"Loading {DATASET.name} ({DATASET.stat().st_size / 1e6:.0f} MB)")
    rows = json.loads(DATASET.read_text())
    print(f"  {len(rows):,} rows in dataset")

    records = {idx: rows[idx] for idx in indices if idx < len(rows)}
    missing = wanted - set(records)
    if missing:
        raise SystemExit(f"{len(missing)} indices missing from dataset, aborting")
    print(f"  resolved all {len(records)} Round-1 indices")
    return indices, records


def build_workbook(ordered, title_note, sheet_name="Verify Reviews"):
    wb = Workbook()
    inst = wb.active
    inst.title = "Instructions"
    lines = [title_note, ""] + INSTRUCTIONS
    for i, line in enumerate(lines, 1):
        inst.cell(row=i, column=1, value=line)
        if i == 1:
            inst.cell(row=i, column=1).font = Font(bold=True, size=14)
    inst.column_dimensions["A"].width = 90

    ws = wb.create_sheet(sheet_name)
    header_fill = PatternFill("solid", fgColor="2E75B6")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(left=Side(style="thin", color="CCCCCC"),
                    right=Side(style="thin", color="CCCCCC"),
                    top=Side(style="thin", color="CCCCCC"),
                    bottom=Side(style="thin", color="CCCCCC"))

    for ci, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill, c.font = header_fill, header_font
        c.alignment = Alignment(horizontal="center")
        c.border = border

    for ri, (idx, r) in enumerate(ordered, start=2):
        row = [idx, r["text"], r.get("rating"), r.get("app_id"),
               r.get("v5_label"), "", "", ""]
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.border = border

    widths = {"row_id": 10, "review_text": 60, "rating": 7, "app_id": 22,
              "predicted_label": 16, "correct_yn": 12,
              "correct_label_if_no": 18, "comments": 30}
    for ci, h in enumerate(HEADERS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(h, 15)
    ws.freeze_panes = "B2"
    return wb


README = """# Round-2 Verification Task - README

Thanks for helping with this annotation round.

## What's in this folder

- `calibration_set_round2.xlsx` - 20 reviews. Do these FIRST. We compare and
  discuss disagreements before the main task starts, so everyone is aligned.
- `annotator_D.xlsx` or `annotator_E.xlsx` - your assigned main task, 490 reviews.
  Fabiha will tell you which letter is yours.
- This README.

## Task in one paragraph

You will see 490 mobile-app reviews. Each one already carries a predicted label
from our classifier. Read the review and decide whether that label is correct (Y)
or wrong (N). If wrong, write the correct label.

## Steps

1. Open the calibration file, read the Instructions sheet, fill in the 20 rows.
2. Send the calibration file back. After we discuss disagreements you are cleared
   to start the main task.
3. Open your `annotator_X.xlsx` and work through the 490 rows.
4. Estimated time: about 3 hours total. Work in chunks, just save the file.

## Ground rules

- Do NOT look at any other annotator's file, and do not discuss individual rows
  while the round is open. We are measuring inter-annotator agreement, so the
  judgements have to be independent.
- Do not re-sort the rows. The row order matches the earlier annotation sheets so
  the supervisor can cross-check the columns side by side.
- Use the exact lowercase label strings listed in the Instructions sheet.
- If you are unsure, mark Y when the label is reasonable. Mark N only when you are
  confident it is wrong.
- Leave a row blank rather than guessing if you truly cannot decide, and note it in
  `comments`. Blank rows are excluded from the analysis.

## Labels

| label | example |
|---|---|
| bug_report | "App keeps crashing when I open it" |
| feature_request | "Please add a dark mode" |
| performance | "Super slow on my phone, takes forever to load" |
| usability | "Hard to find the settings menu" |
| compatibility | "Doesn't work on Samsung Galaxy S22" |
| praise | "Best app ever, love it!" |
| other | "Hi, just downloaded this" |

## When you are done

Save the file under its original name and send it back. Do not rename it, and do
not save it over anyone else's copy.
"""


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    indices, records = load_selected_records()

    key = json.loads(MASTER_KEY.read_text())
    cal_indices = [int(i) for i in key["calibration_indices"]]
    print(f"Loading dataset rows for {len(cal_indices)} calibration reviews")
    all_rows = json.loads(DATASET.read_text())
    cal = [(i, all_rows[i]) for i in cal_indices]

    wb = build_workbook(cal, "CALIBRATION ROUND 2 - 20 reviews. Both annotators do "
                             "the SAME 20 first, to align on the categories.",
                        sheet_name="Calibration")
    wb.save(OUT_DIR / "calibration_set_round2.xlsx")
    print(f"  saved {OUT_DIR / 'calibration_set_round2.xlsx'}")

    order_log = {}
    for code, display in ANNOTATORS.items():
        ordered = [(i, records[i]) for i in indices]
        wb = build_workbook(ordered,
                            f"VERIFICATION TASK - {display}. Do NOT consult the other "
                            f"annotator or any earlier sheet while filling this in.")
        path = OUT_DIR / f"annotator_{code}.xlsx"
        wb.save(path)
        order_log[code] = {"row_order": "round1_order"}
        print(f"  saved {path}  ({len(ordered)} rows, blank annotation columns)")

    (OUT_DIR / "round2_key.json").write_text(json.dumps({
        "source_indices": indices,
        "calibration_indices": cal_indices,
        "annotator_orders": order_log,
        "note": "Round-2 independent re-annotation of the Round-1 490-review sample. "
                "All sheets use the Round-1 row order so they line up for manual "
                "cross-checking. Align in code by row_id, never by row position.",
    }, indent=2))
    print(f"  saved {OUT_DIR / 'round2_key.json'}  (keep private)")

    (OUT_DIR / "00_README.md").write_text(README)
    print(f"  saved {OUT_DIR / '00_README.md'}")

    print("\nSend each annotator: 00_README.md, calibration_set_round2.xlsx, "
          "and their own annotator_X.xlsx. Keep round2_key.json to yourself.")


if __name__ == "__main__":
    main()
