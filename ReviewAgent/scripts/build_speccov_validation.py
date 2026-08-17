"""
Build the SpecCov human-validation sheets.

Why this exists
---------------
The paper releases SpecCov as an extractive-coverage faithfulness score, but its only
supporting number came from an automatic rubric that uses the same overlap procedure, so
SpecCov was effectively validated against itself. This task collects independent human
faithfulness judgements on the same specs, so the two can be correlated.

Design note that matters for validity
-------------------------------------
SpecCov scores a spec against `representative_reviews` + `first_5_review_texts` + the
cluster's `auto_name`. The sheet therefore shows a rater exactly that same evidence, in
that order. An earlier draft of this sheet showed only `first_5_review_texts`, which would
have made raters mark supported claims as unsupported whenever the supporting sentence sat
in the other list.

60 specs, 20 each from three generators (taxonomy-grounded LLM, free-form LLM, lead-author
reference), shuffled and unlabelled. Three sheets, one per rater.

Output (annotator_materials/speccov_validation/)
    speccov_validation_A.xlsx, _D.xlsx, _E.xlsx
    speccov_key.json     condition + SpecCov score per spec; keep private
    00_README.md

Usage
    python3 scripts/build_speccov_validation.py
"""

import json
import random
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "scripts"))
from speccov import speccov_score  # noqa: E402

SPECS = REPO / "data/processed/issue_specs"
OUT = REPO / "annotator_materials/speccov_validation"
CONDITIONS = {"llm_taxonomy": "specs_with_taxonomy.json",
              "llm_free_form": "specs_free_form.json",
              "human_ref": "specs_human_written.json"}
PER_CONDITION = 20
SEED = 7
RATERS = ["A", "D", "E"]

HEAD = ["spec_uid", "issue_type", "cluster_topic", "source_reviews", "issue_spec",
        "faithfulness_1_to_5", "unsupported_details", "notes"]

INSTRUCTIONS = [
    "SPEC FAITHFULNESS RATING TASK",
    "",
    "Each row shows the source reviews a specification was written from, and the",
    "specification itself. Judge ONE thing: is the spec grounded in those reviews?",
    "",
    "The source_reviews cell contains every review the spec writer had access to.",
    "If a claim is not traceable to that cell, it is unsupported, full stop.",
    "",
    "faithfulness_1_to_5",
    "  5  every claim in the spec is supported by the source reviews",
    "  4  supported, with at most one minor unsupported detail",
    "  3  mostly supported, some details not traceable to the reviews",
    "  2  several claims not supported by the reviews",
    "  1  the spec describes something the reviews do not say",
    "",
    "unsupported_details",
    "  List every concrete item the spec asserts that the reviews never mention:",
    "  a device name, an OS version, a component, a reproduction step, a number.",
    "  Write 'none' if there are none. Separate items with a semicolon.",
    "  This column is the important one. It is what tells us whether an automatic",
    "  grounding score detects the same thing a human calls a hallucination.",
    "",
    "Rules of thumb",
    "  - Rate grounding, not writing quality. A clumsy but fully grounded spec is a 5.",
    "  - A polished spec that invents a device model is a 2.",
    "  - Generalising is fine: 'users report crashes' when three reviews say 'it crashes'",
    "    is supported. Inventing specifics is not: 'crashes on Android 12' when no review",
    "    mentions a version is unsupported.",
    "  - Reasonable inference from the reviews counts as supported. A named entity that",
    "    appears nowhere in the reviews does not, however plausible it sounds.",
    "  - Judge the spec as a whole. One invented field in an otherwise grounded spec is a 4.",
    "",
    "Specs come from three different writers, shuffled and unlabelled. Do not try to work",
    "out which is which, and do not discuss rows with the other raters while the round is",
    "open. About 45 minutes. Save as you go.",
    "",
]

README = """# SpecCov Validation Task - README

One file per rater: `speccov_validation_A.xlsx`, `_D.xlsx`, `_E.xlsx`. Fabiha will tell you
which is yours. All three contain the same 60 rows in the same order.

## What you are judging

Each row gives you the source reviews a specification was written from, and the
specification. You judge whether the spec is *grounded* in those reviews. You are not
judging whether the spec is well written, well formatted, or useful.

Fill two columns:

- `faithfulness_1_to_5` - 5 means everything in the spec traces back to the reviews,
  1 means the spec describes something the reviews do not say.
- `unsupported_details` - list anything the spec asserts that the reviews never mention.
  Write `none` if there is nothing. This column is the point of the whole task.

The Instructions sheet inside the file has the full scale and the rules of thumb. Read it
before starting.

## Why we are asking

We built an automatic score that tries to detect ungrounded specifications. We do not yet
know whether it detects what a person would call a hallucination. Your `unsupported_details`
column is the ground truth we compare it against, so guessing hurts more than leaving a row
blank. If you cannot decide, leave the row blank and say why in `notes`.

## Ground rules

- Do not discuss individual rows with the other raters while the round is open.
- Do not re-sort the rows.
- Save the file under its original name and send it back.

About 45 minutes.
"""


def render_reviews(cluster):
    """Exactly the evidence SpecCov scores against, in the same order."""
    lines = []
    rep = cluster.get("representative_reviews") or []
    first5 = cluster.get("first_5_review_texts") or []
    for r in rep:
        lines.append(f"- {r}")
    for r in first5:
        if r not in rep:
            lines.append(f"- {r}")
    return "\n".join(lines)


def render_spec(spec):
    skip = {"issue_id", "cluster_id", "condition", "model", "spec_raw", "spec_json"}
    return "\n".join(f"{k}: {v}" for k, v in spec.items() if k not in skip and v)


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    clusters = {c["cluster_id"]: c for c in json.load(open(SPECS / "sample_100_clusters.json"))}
    rng = random.Random(SEED)

    rows = []
    for cond, filename in CONDITIONS.items():
        specs = json.load(open(SPECS / filename))
        pool = [s for s in specs if s.get("cluster_id") in clusters]
        for s in rng.sample(pool, min(PER_CONDITION, len(pool))):
            cl = clusters[s["cluster_id"]]
            rows.append({
                "spec_uid": f"{cond}:{s['cluster_id']}",
                "issue_type": s.get("issue_type", cl.get("issue_type", "")),
                "cluster_topic": cl.get("auto_name", ""),
                "source_reviews": render_reviews(cl),
                "issue_spec": render_spec(s),
                "_condition": cond,
                "_speccov": speccov_score(s, cl),
            })
    rng.shuffle(rows)
    print(f"{len(rows)} specs, {PER_CONDITION} per condition", file=sys.stderr)

    widths = {"spec_uid": 22, "issue_type": 15, "cluster_topic": 26, "source_reviews": 62,
              "issue_spec": 72, "faithfulness_1_to_5": 18, "unsupported_details": 34, "notes": 22}
    for code in RATERS:
        wb = Workbook()
        ins = wb.active
        ins.title = "Instructions"
        lines = [f"SPEC FAITHFULNESS RATING - Rater {code}. Do NOT consult the other raters.",
                 ""] + INSTRUCTIONS
        for i, line in enumerate(lines, 1):
            ins.cell(row=i, column=1, value=line)
            if i == 1:
                ins.cell(row=i, column=1).font = Font(bold=True, size=14)
        ins.column_dimensions["A"].width = 95

        ws = wb.create_sheet("Specs")
        fill = PatternFill("solid", fgColor="2E75B6")
        font = Font(bold=True, color="FFFFFF")
        border = Border(*[Side(style="thin", color="CCCCCC")] * 4)
        for ci, h in enumerate(HEAD, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.fill, c.font = fill, font
            c.alignment = Alignment(horizontal="center")
            c.border = border
        for ri, r in enumerate(rows, start=2):
            for ci, h in enumerate(HEAD, 1):
                c = ws.cell(row=ri, column=ci, value=r.get(h, ""))
                c.alignment = Alignment(wrap_text=True, vertical="top")
                c.border = border
        for ci, h in enumerate(HEAD, 1):
            ws.column_dimensions[get_column_letter(ci)].width = widths.get(h, 15)
        ws.freeze_panes = "C2"
        path = OUT / f"speccov_validation_{code}.xlsx"
        wb.save(path)
        print(f"  saved {path}", file=sys.stderr)

    (OUT / "speccov_key.json").write_text(json.dumps(
        {r["spec_uid"]: {"condition": r["_condition"], "speccov": r["_speccov"]} for r in rows},
        indent=2))
    (OUT / "00_README.md").write_text(README)
    print(f"  saved key and README in {OUT}", file=sys.stderr)


if __name__ == "__main__":
    main()
