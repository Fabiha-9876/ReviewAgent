"""
Score the five-dimension rubric on the full 100-spec set, not a 28-spec subsample.

The headline rubric number in the paper came from a stratified 28-spec subsample, chosen
when judge throughput on this machine was the binding constraint. A reviewer reasonably
asked whether 28 specs support a claim about corpus-scale behaviour. All 100 taxonomy
specs exist, so this run scores every one of them, plus the 20 lead-author reference specs
used for judge calibration.

Judge and rubric are unchanged from the 28-spec run, so the two are directly comparable and
the difference between them measures subsample error rather than a change in method.

Outputs
    data/processed/ablations/qwen_judge_5dim_rubric_full.json
    data/processed/ablations/qwen_judge_5dim_rubric_full.txt

Usage
    python3 scripts/run_full_rubric_100.py
"""

from __future__ import annotations

import json
import re
import statistics
import sys
import time
from collections import defaultdict
from pathlib import Path

import openpyxl
import torch
from transformers import AutoModelForCausalLM, AutoTokenizer

REPO = Path(__file__).resolve().parent.parent
SPECS = REPO / "data/processed/issue_specs"
REF_XLSX = REPO / "human_work/reference_specs.xlsx"
OUT = REPO / "data/processed/ablations/qwen_judge_5dim_rubric_full.json"
OUT_TXT = OUT.with_suffix(".txt")
JUDGE = "Qwen/Qwen2.5-3B-Instruct"
DIMS = ["completeness", "accuracy", "actionability", "specificity", "clarity"]

SYS = """You are an expert software-engineering reviewer scoring a generated bug or feature issue specification on five dimensions, each on a 1-5 scale where 5 is best.

Dimensions:
1. Completeness: are all expected fields populated with substantive content?
2. Accuracy: does the spec faithfully reflect the source review evidence?
3. Actionability: can a developer act on this without going back to the user?
4. Specificity: is it specific enough to file in a defect tracker?
5. Clarity: is it clearly written and well-structured?

Output EXACTLY in this format (one line per dimension):
completeness: <1-5>
accuracy: <1-5>
actionability: <1-5>
specificity: <1-5>
clarity: <1-5>"""

FIELDS = ("title", "issue_type", "description", "severity", "affected_component",
          "steps_to_reproduce", "expected_behavior", "actual_behavior",
          "user_story", "acceptance_criteria", "nfr_category", "nielsen_heuristic",
          "device_os_matrix")


def spec_to_text(spec):
    parts = []
    for k in FIELDS:
        v = spec.get(k)
        if not v:
            continue
        if isinstance(v, list):
            v = "; ".join(str(x)[:80] for x in v[:5])
        elif isinstance(v, dict):
            v = str(v)[:200]
        parts.append(f"{k}: {v}")
    return "\n".join(parts)


def load_reference_specs():
    """The 20 lead-author reference specs, from the sheet they were written in."""
    ws = openpyxl.load_workbook(REF_XLSX)["Specs"]
    hdr = [c.value for c in ws[1]]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        d = dict(zip(hdr, row))
        spec = {k[len("your_"):]: v for k, v in d.items()
                if isinstance(k, str) and k.startswith("your_") and v}
        if spec:
            spec["issue_type"] = d.get("issue_type")
            spec["cluster_id"] = d.get("cluster_id")
            out.append(spec)
    return out


def main():
    taxonomy = json.load(open(SPECS / "specs_with_taxonomy.json"))
    reference = load_reference_specs()
    print(f"{len(taxonomy)} taxonomy specs, {len(reference)} reference specs", file=sys.stderr)

    device = "mps" if torch.backends.mps.is_available() else "cpu"
    print(f"Loading {JUDGE} on {device}", file=sys.stderr)
    tok = AutoTokenizer.from_pretrained(JUDGE)
    model = AutoModelForCausalLM.from_pretrained(
        JUDGE, torch_dtype=torch.float16 if device != "cpu" else torch.float32
    ).to(device).eval()

    def judge(text):
        msgs = [{"role": "system", "content": SYS},
                {"role": "user",
                 "content": f"IssueSpec to score:\n{text[:2000]}\n\nScore on the 5 dimensions:"}]
        chat = tok.apply_chat_template(msgs, tokenize=False, add_generation_prompt=True)
        enc = tok(chat, return_tensors="pt").to(device)
        with torch.inference_mode():
            gen = model.generate(**enc, max_new_tokens=80, do_sample=False,
                                 pad_token_id=tok.eos_token_id)
        text_out = tok.decode(gen[0, enc["input_ids"].shape[1]:], skip_special_tokens=True)
        scores = {}
        for line in text_out.splitlines():
            m = re.match(r"\s*(completeness|accuracy|actionability|specificity|clarity)"
                         r"\s*:\s*([1-5])", line.lower())
            if m:
                scores[m.group(1)] = int(m.group(2))
        return scores

    def score_set(specs, tag):
        rows, t0 = [], time.time()
        for i, s in enumerate(specs, 1):
            body = spec_to_text(s)
            if len(body) < 50:
                continue
            sc = judge(body)
            if len(sc) < len(DIMS):
                continue
            rows.append({"cluster_id": s.get("cluster_id"),
                         "issue_type": s.get("issue_type"), **sc})
            print(f"  [{tag}] {i}/{len(specs)}  ({time.time() - t0:.0f}s)",
                  file=sys.stderr, end="\r")
        print(f"\n  [{tag}] scored {len(rows)} in {time.time() - t0:.0f}s", file=sys.stderr)
        return rows

    tax_rows = score_set(taxonomy, "taxonomy")
    ref_rows = score_set(reference, "reference")

    def summarise(rows):
        per_spec = [statistics.mean(r[d] for d in DIMS) for r in rows]
        return {"n": len(rows),
                "overall_mean": round(statistics.mean(per_spec), 3),
                "overall_sd": round(statistics.pstdev(per_spec), 3),
                "per_dim": {d: round(statistics.mean(r[d] for r in rows), 2) for d in DIMS}}

    by_type = defaultdict(list)
    for r in tax_rows:
        by_type[r["issue_type"]].append(r)

    report = {
        "judge": JUDGE, "rubric_dimensions": DIMS,
        "taxonomy_full": summarise(tax_rows),
        "reference_full": summarise(ref_rows),
        "taxonomy_per_issue_type": {t: summarise(rs) for t, rs in sorted(by_type.items())},
        "subsample_comparison": {
            "previous_n": 28, "previous_mean": 3.89,
            "note": "same judge and rubric; the difference measures subsample error",
        },
        "rows": {"taxonomy": tax_rows, "reference": ref_rows},
    }
    report["subsample_comparison"]["full_mean"] = report["taxonomy_full"]["overall_mean"]
    report["subsample_comparison"]["delta"] = round(
        report["taxonomy_full"]["overall_mean"] - 3.89, 3)
    OUT.write_text(json.dumps(report, indent=2))

    L = ["=" * 74, "FIVE-DIMENSION RUBRIC ON THE FULL SPEC SET", "=" * 74, "",
         f"judge: {JUDGE}", "",
         f"  {'set':16s}{'n':>5s}{'mean':>7s}{'sd':>7s}"
         + "".join(f"{d[:5]:>8s}" for d in DIMS)]
    for name, key in (("taxonomy", "taxonomy_full"), ("lead-author ref", "reference_full")):
        v = report[key]
        L.append(f"  {name:16s}{v['n']:>5d}{v['overall_mean']:>7.2f}{v['overall_sd']:>7.2f}"
                 + "".join(f"{v['per_dim'][d]:>8.2f}" for d in DIMS))
    L += ["", "  per issue type (taxonomy condition)", "  " + "-" * 60]
    for t, v in report["taxonomy_per_issue_type"].items():
        L.append(f"  {str(t):18s} n={v['n']:<4d} mean={v['overall_mean']:.2f}")
    sc = report["subsample_comparison"]
    L += ["", f"  28-spec subsample: {sc['previous_mean']:.2f}   "
              f"full {report['taxonomy_full']['n']}-spec: {sc['full_mean']:.2f}   "
              f"delta {sc['delta']:+.2f}", ""]
    OUT_TXT.write_text("\n".join(L) + "\n")
    print("\n" + "\n".join(L))


if __name__ == "__main__":
    main()
